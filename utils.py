import pandas as pd
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font, Alignment

def record_to_dict(
    age: int,
    marital_status: str,
    education: str,
    occupation: str,
    family_size: int,
    region: str,
    income: float,
    actual_expenses: float,
    housing_type: str,
    electricity_bill: float
) -> dict:
    return {
        "age": age,
        "marital_status": marital_status,
        "education": education,
        "occupation": occupation,
        "family_size": family_size,
        "region": region,
        "income": income,
        "actual_expenses": actual_expenses,
        "housing_type": housing_type,
        "electricity_bill": electricity_bill,
        "timestamp": datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    }

def get_trust_color(score: int) -> str:
    if score >= 75:
        return "green"
    elif score >= 50:
        return "orange"
    else:
        return "red"

def get_status_emoji(status: str) -> str:
    if "موثوق" in status:
        return "✅"
    elif "مراجعة" in status:
        return "⚠️"
    else:
        return "❌"

def format_issues(issues: list) -> str:
    if not issues:
        return "لا توجد مشاكل"
    return "\n".join([f"• {issue}" for issue in issues])

def csv_to_records(df: pd.DataFrame) -> list:
    required_columns = [
        'age', 'marital_status', 'education', 'occupation',
        'family_size', 'region', 'income', 'actual_expenses',
        'housing_type', 'electricity_bill'
    ]
    missing = [col for col in required_columns if col not in df.columns]
    if missing:
        raise ValueError(f"الأعمدة الناقصة في الملف: {missing}")
    return df[required_columns].to_dict(orient='records')

def results_to_dataframe(results: list) -> pd.DataFrame:
    rows = []
    for result in results:
        manager = result.get('manager', {})
        demo = result.get('demographic', {})
        financial = result.get('financial', {})
        original = result.get('original_record', {})
        rows.append({
            'رقم السجل': result.get('record_index', ''),
            'المنطقة': original.get('region', ''),
            'الدخل': original.get('income', ''),
            'الإنفاق الفعلي': original.get('actual_expenses', ''),
            'نوع السكن': original.get('housing_type', ''),
            'درجة الموثوقية': manager.get('trust_score', 0),
            'الحالة النهائية': manager.get('status', ''),
            'التوصية': manager.get('recommendation', ''),
            'مشاكل ديموغرافية': format_issues(demo.get('issues', [])),
            'مشاكل مالية': format_issues(financial.get('issues', [])),
            'وقت التحليل': datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        })
    return pd.DataFrame(rows)

def export_to_excel(results: list, filename: str = None) -> str:
    if filename is None:
        filename = f"eco_guard_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    df = results_to_dataframe(results)
    df.to_excel(filename, index=False, engine='openpyxl')

    # ===== تلوين الصفوف حسب الحالة =====
    wb = load_workbook(filename)
    ws = wb.active

    # ألوان الحالات
    GREEN_FILL  = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # موثوق
    YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # يحتاج مراجعة
    RED_FILL    = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # مرفوض

    # تلوين رأس الجدول
    HEADER_FILL = PatternFill(start_color="1A6B52", end_color="1A6B52", fill_type="solid")
    for cell in ws[1]:
        cell.fill = HEADER_FILL
        cell.font = Font(bold=True, color="FFFFFF")
        cell.alignment = Alignment(horizontal="center")

    # تحديد عمود الحالة النهائية
    status_col = None
    for col in ws.iter_cols(1, ws.max_column, 1, 1):
        if col[0].value == "الحالة النهائية":
            status_col = col[0].column
            break

    # تلوين كل صف حسب الحالة
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if status_col:
            status_value = row[status_col - 1].value or ""
            if "موثوق" in status_value and "مراجعة" not in status_value:
                fill = GREEN_FILL
            elif "مراجعة" in status_value:
                fill = YELLOW_FILL
            else:
                fill = RED_FILL
            for cell in row:
                cell.fill = fill

    # ضبط عرض الأعمدة تلقائياً
    for col in ws.columns:
        max_length = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[col[0].column_letter].width = min(max_length + 4, 40)

    wb.save(filename)
    return filename